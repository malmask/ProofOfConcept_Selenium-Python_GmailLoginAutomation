import openpyxl

class ExcelHandler:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path)
        self.sheet = self.workbook.active

    def read_credentials(self):
        credentials = []
        for row in self.sheet.iter_rows(min_row=2, values_only=True):  
            credentials.append({
                'login_id': row[0],
                'password': row[1],
            })
        return credentials

    def read_email_details(self):
        email_details = {
            'to': self.sheet.cell(row=2, column=3).value,
            'subject': self.sheet.cell(row=2, column=4).value,
            'body': self.sheet.cell(row=2, column=5).value,
            'attachment_path': self.sheet.cell(row=2, column=6).value,
        }
        return email_details

    def update_status(self, row, status, message):
        self.sheet.cell(row=row, column=7, value=status)
        self.sheet.cell(row=row, column=8, value=message)
        self.workbook.save(self.file_path)
